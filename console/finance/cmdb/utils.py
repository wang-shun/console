# coding=utf-8
from openpyxl import load_workbook
from rest_framework.utils import model_meta

from console.common import serializers

EXCELS = {
    'db': ['cfg_id', 'name', 'version', 'memo', 'instance', 'net', 'capacity'],
    'sys': ['cfg_id', 'name', 'version', 'hosts', 'man', 'weight'],
}


def parse_excel(filename, tpe):
    data = list()
    fields = EXCELS.get(tpe)
    if fields:
        excel = load_workbook(filename=filename)
        ws = excel.get_sheet_names()
        table = excel.get_sheet_by_name(ws[0])
        started = False
        for i in range(1, table.max_row):
            cfg_id = table.cell(row=i, column=1).value
            if not started:
                started = (u'编号' == cfg_id)
                continue
            elif not cfg_id:
                break
            else:
                data.append({
                    field: table.cell(row=i, column=j).value
                    for j, field in enumerate(fields, 1)
                })
    return data


def get_default_field_names(model):
    """ 获取一个model的REST Framework默认的渲染field names """
    field_info = model_meta.get_field_info(model)
    return serializers.ModelSerializer().get_default_field_names({}, field_info)
