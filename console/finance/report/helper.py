# coding=utf-8
import copy
import datetime
import functools
from collections import OrderedDict

from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Font, borders, colors
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from dateutil.relativedelta import relativedelta
from django.utils import timezone

from console.common.zones.models import ZoneModel
from console.finance.tickets.models import FinanceTicketModel
from console.finance.report.models import (
    PhysicalMachineUseRecord,
    VirtualResourceSnapshot,
    VirtualMachineUseRecord,
    BusinessMonitorSnapshot,
)


def list_cal(l1, l2, operation="-", in_place=False, null_as=None):
    # print "l1:", l1
    # print "l2:", l2
    if not isinstance(l2, (list, tuple)):
        l2 = [l2] * len(l1)
    operation_dict = {
        "-": lambda a, b: a - b,
        "+": lambda a, b: a + b,
        "*": lambda a, b: a * b,
        "/": lambda a, b: a * 1.0 / b if b != 0 else 0,
    }

    def null_wrap(func):
        @functools.wraps(func)
        def _func(a, b, null_as=null_as):
            if a is None or b is None:
                if null_as == "return_zero":
                    return 0
                elif null_as == "return_null":
                    return None
                if a is None:
                    a = null_as
                if b is None:
                    b = null_as
            return func(a, b)
        return _func

    operate_func = null_wrap(operation_dict[operation])
    if in_place:
        for index in range(len(l1)):
            l1[index] = operate_func(l1[index], l2[index], null_as=null_as)
    else:
        return map(operate_func, l1, l2, [null_as] * len(l1))


def list_avg(l, null=0):
    """ 求列表平均值，自动过滤None，空列表返回null """
    filter_l = filter(lambda v: v is not None, l)
    if filter_l:
        return sum(filter_l) / float(len(filter_l))
    else:
        return null


def merge_value(values, report_type):
    if report_type == "day":  # 不需要merge
        return values
    new_values = []
    for i in range(0, len(values), 24):  # 将24小时的数据合并为一天
        new_values.append(list_avg(values[i: (i + 24)], null=None))
    return new_values


def auto_replace_none_value(l):
    """ 检测列表中为None的值，进行替换 """
    l_len = len(l)
    new_l = []
    for index, value in enumerate(l):
        if value is None:
            pre_value = l[index - 1] if index > 0 and l[index - 1] is not None else None
            next_value = l[index + 1] if index < l_len - 1 and l[index + 1] is not None else None
            if pre_value is None and next_value is not None:
                value = (pre_value + next_value) / 2.0
            elif pre_value is not None:
                value = pre_value
            elif next_value is not None:
                value = next_value
            else:
                value = 0
        new_l.append(value)
    return new_l


def result_precess(value):
    if value is None:
        return 0
    return int(round(value, 0))


def safe_divide(a, b):
    if a is None or b is None or b == 0:
        return 0
    return a * 1.0 / b


def get_report_end(start, report_type):
    if report_type == "day":
        end = start + datetime.timedelta(days=1)
    elif report_type == "week":
        end = start + datetime.timedelta(days=7 - start.weekday())
    else:
        tmp = start + relativedelta(months=1)
        end = datetime.datetime(tmp.year, tmp.month, 1)
    return end


def get_ticket_report(request, type_, datetime_start, datetime_end, zone=None, **kwargs):
    zone = zone or request.data.get("zone")
    zone = ZoneModel.objects.filter(name=zone).first()
    tickets = FinanceTicketModel.objects.filter(
        commit_time__gte=datetime_start, commit_time__lt=datetime_end, zone=zone
    )
    ticket_dict = {}
    datetime_end = timezone.make_aware(datetime_end, timezone.get_default_timezone())  # 否则下面时间比较会报错
    for ticket in tickets:
        name = ticket.ticket_type.ticket_name
        if name not in ticket_dict:
            ticket_dict[name] = {
                "name": name,
                "total": 0,
                "finished": 0,
            }
        ticket_dict[name]["total"] += 1
        if ticket.cur_status == "finish" and ticket.last_update_time < datetime_end:
            ticket_dict[name]["finished"] += 1
    tickets = ticket_dict.values()
    tickets.sort(key=lambda v: v["total"])
    ticket_info = {
        "names": [],
        "totals": [],
        "finish_nums": [],
        "finish_rates": [],
    }
    finish_rate_low_tickets = []
    ticket_num = 0
    for t in tickets:
        ticket_info["names"].append(t["name"])
        ticket_info["totals"].append(t["total"])
        ticket_info["finish_nums"].append(t["finished"])
        ticket_num += t["total"]
        finish_rate = t["finished"] * 100.0 / t["total"] if t["total"] else 100
        if finish_rate < 60:
            finish_rate_low_tickets.append(t["name"])
        ticket_info["finish_rates"].append(result_precess(finish_rate))
    return [{
        "ticket_num": ticket_num,
        "finish_rate_low_tickets": finish_rate_low_tickets,
        "ticket": ticket_info,
    }]


def get_physical_resource_report(request, type_, datetime_start, datetime_end, zone=None, **kwargs):
    timestamp_start = kwargs["time_start"]
    zone = zone or request.data.get("zone")
    zone = ZoneModel.objects.filter(name=zone).first()
    time_str_start = datetime_start.strftime("%Y%m%d00")
    time_str_end = datetime_end.strftime("%Y%m%d00")
    records = PhysicalMachineUseRecord.objects.filter(time__gte=time_str_start, time__lt=time_str_end, zone=zone)
    expect_point_num = (datetime_end - datetime_start).days * 24
    init_value_list = [None] * expect_point_num
    data_dict = {}

    def add_value(item_dict, index, record):
        items = ["cpu_total", "cpu_used", "memory_total", "memory_used", "disk_total", "disk_used"]
        for item in items:
            record_value = getattr(record, item)
            if record_value is not None:
                if item_dict[item][index] is None:
                    item_dict[item][index] = 0
                item_dict[item][index] += record_value

    for record in records:
        cabinet = record.cabinet
        if cabinet not in data_dict:
            data_dict[cabinet] = {
                "cpu_total": copy.deepcopy(init_value_list),
                "cpu_used": copy.deepcopy(init_value_list),
                "memory_total": copy.deepcopy(init_value_list),
                "memory_used": copy.deepcopy(init_value_list),
                "disk_total": copy.deepcopy(init_value_list),
                "disk_used": copy.deepcopy(init_value_list),
            }
        time_datetime = datetime.datetime.strptime(record.time, "%Y%m%d%H")
        timedelta = time_datetime - datetime_start
        index = timedelta.days * 24 + timedelta.seconds / 60 / 60
        add_value(data_dict[cabinet], index, record)
    # logger.info(data_dict)

    for item_dict in data_dict.values():
        for item, values in item_dict.items():
            item_dict[item] = merge_value(values, type_)

    cabinet_use_rate = {
        "names": [],
        "cpu_use_rates": [],
        "memory_use_rates": [],
        "disk_use_rates": [],
    }
    cabinets = sorted(data_dict.keys())
    for cabinet in cabinets:
        item_dict = data_dict[cabinet]
        cpu_total = list_avg(item_dict["cpu_total"])
        cpu_used = list_avg(item_dict["cpu_used"])
        memory_total = list_avg(item_dict["memory_total"])
        memory_used = list_avg(item_dict["memory_used"])
        disk_total = list_avg(item_dict["disk_total"])
        disk_used = list_avg(item_dict["disk_used"])
        cabinet_use_rate["names"].append(cabinet)
        cabinet_use_rate["cpu_use_rates"].append(result_precess(safe_divide(cpu_used, cpu_total) * 100))
        cabinet_use_rate["memory_use_rates"].append(result_precess(safe_divide(memory_used, memory_total) * 100))
        cabinet_use_rate["disk_use_rates"].append(result_precess(safe_divide(disk_used, disk_total) * 100))

    high_load_cabinet = {
        "cpu": [],
        "memory": [],
        "disk": [],
    }
    for i, name in enumerate(cabinet_use_rate["names"]):
        if cabinet_use_rate["cpu_use_rates"][i] >= 80:
            high_load_cabinet["cpu"].append(name)
        if cabinet_use_rate["memory_use_rates"][i] >= 80:
            high_load_cabinet["memory"].append(name)
        if cabinet_use_rate["disk_use_rates"][i] >= 80:
            high_load_cabinet["disk"].append(name)

    if type_ != "day":
        expect_point_num = expect_point_num / 24
    item_dict = {
        "cpu_total": [0] * expect_point_num,
        "cpu_used": [0] * expect_point_num,
        "memory_total": [0] * expect_point_num,
        "memory_used": [0] * expect_point_num,
        "disk_total": [0] * expect_point_num,
        "disk_used": [0] * expect_point_num,
    }
    for cabinet_item_dict in data_dict.values():
        for key in item_dict.keys():
            list_cal(item_dict[key], cabinet_item_dict[key], "+", in_place=True, null_as=0)
    load_trend = {
        "times": [],
        "cpu_use_rates":
            map(result_precess,
                list_cal(list_cal(item_dict["cpu_used"], item_dict["cpu_total"], "/", null_as=0), 100, "*")),
        "memory_use_rates":
            map(result_precess,
                list_cal(list_cal(item_dict["memory_used"], item_dict["memory_total"], "/", null_as=0), 100, "*")),
        "disk_use_rates":
            map(result_precess,
                list_cal(list_cal(item_dict["disk_used"], item_dict["disk_total"], "/", null_as=0), 100, "*")),
    }
    interval = 3600 if type_ == "day" else 24 * 3600
    for i in range(expect_point_num):
        load_trend["times"].append(timestamp_start + i * interval)

    return [{
        "high_load_cabinet": high_load_cabinet,
        "cabinet_use_rate": cabinet_use_rate,
        "load_trend": load_trend,
    }]


def get_virtual_resource_report(request, type_, datetime_start, datetime_end, zone=None, **kwargs):
    timestamp_start = kwargs["time_start"]
    zone = zone or request.data.get("zone")
    zone = ZoneModel.objects.filter(name=zone).first()
    time_str_start = datetime_start.strftime("%Y%m%d00")
    time_str_end = datetime_end.strftime("%Y%m%d00")

    records = VirtualMachineUseRecord.objects.filter(time__gte=time_str_start, time__lt=time_str_end, zone=zone)
    expect_point_num = (datetime_end - datetime_start).days * 24
    init_value_list = [None] * expect_point_num
    items = ["cpu_total", "cpu_used", "memory_total", "memory_used", "disk_total", "disk_used"]
    item_dict = {item: copy.deepcopy(init_value_list) for item in items}
    for record in records:
        time_datetime = datetime.datetime.strptime(record.time, "%Y%m%d%H")
        timedelta = time_datetime - datetime_start
        index = timedelta.days * 24 + timedelta.seconds / 60 / 60
        for item_index in range(0, len(items), 2):
            item1 = items[item_index]
            item2 = items[item_index + 1]
            value1 = getattr(record, item1)
            value2 = getattr(record, item2)
            if value1 and value2:  # 只有total和used都有
                item_dict[item1][index] = (item_dict[item1][index] or 0) + value1
                item_dict[item2][index] = (item_dict[item2][index] or 0) + value2
    # logger.info("item_dict: %s " % item_dict)
    load_trend = {
        "cpu_use_rates": list_cal(
            list_cal(item_dict["cpu_used"], item_dict["cpu_total"], "/", null_as="return_zero"),
            100, "*"
        ),
        "memory_use_rates": list_cal(
            list_cal(item_dict["memory_used"], item_dict["memory_total"], "/", null_as="return_zero"),
            100, "*"
        ),
        "disk_use_rates": list_cal(
            list_cal(item_dict["disk_used"], item_dict["disk_total"], "/", null_as="return_zero"),
            100, "*"
        ),
    }
    # logger.info("load_trend: %s " % load_trend)
    for item, values in load_trend.items():
        load_trend[item] = map(result_precess, merge_value(values, type_))
    load_trend["times"] = []
    interval = 3600 if type_ == "day" else 24 * 3600
    if type_ != "day":
        expect_point_num = expect_point_num / 24
    for i in range(expect_point_num):
        load_trend["times"].append(timestamp_start + i * interval)

    vr_before_record = VirtualResourceSnapshot.objects.filter(time=time_str_start, zone=zone).first()
    vr_after_record = VirtualResourceSnapshot.objects.filter(time=time_str_end, zone=zone).first()
    name_item_dict = OrderedDict([
        (u"主机", ("vm_num", u"台", int)),
        ("CPU", ("cpu_num", u"核", int)),
        (u"内存", ("memory_total", "G", int)),
        (u"硬盘", ("disk_total", "T", lambda v: round(v / 1024.0, 2))),
        (u"中间件-LB", ("lb_num", u"套", int)),
        (u"中间件-WAF", ("waf_num", u"套", int)),
    ])
    change_dict = {
        "names": [],
        "rate_changes": [],
        "num_changes": [],
        "raw_nums": [],
        "rate_basics": [],
    }
    if vr_before_record and vr_after_record:
        for name, item in name_item_dict.items():
            change_dict["names"].append(name)
            item_name, item_unit, item_convert_func = item
            before_value = getattr(vr_before_record, item_name, None)
            after_value = getattr(vr_after_record, item_name, None)
            if before_value is not None and after_value is not None:
                if before_value:
                    rate_change = int(after_value * 100.0 / before_value - 100)
                    rate_basic = 100
                else:
                    rate_change = 100 if after_value else 0
                    rate_basic = 0
                num_change = "%s%s" % (item_convert_func(after_value - before_value), item_unit)
            else:
                rate_change = 0
                num_change = "0%s" % item_unit
                rate_basic = 100
            change_dict["raw_nums"].append(before_value)
            change_dict["rate_changes"].append(rate_change)
            change_dict["num_changes"].append(num_change)
            change_dict["rate_basics"].append(rate_basic)
    return [{
        "change": change_dict,
        "load_trend": load_trend,
    }]


def get_business_report(request, type_, datetime_start, datetime_end, zone=None, **kwargs):
    zone = zone or request.data.get("zone")
    zone = ZoneModel.objects.filter(name=zone).first()
    time_str_start = datetime_start.strftime("%Y%m%d00")
    time_str_end = datetime_end.strftime("%Y%m%d00")

    records = BusinessMonitorSnapshot.objects.filter(
        time=time_str_end, zone=zone, app_system__isnull=False
    ).select_related("app_system")
    system_dict = {}
    high_risk_systems = []
    for record in records:
        if record.app_system is None:
            continue
        if record.os_leak_num + record.site_leak_num + record.weak_order_num + record.horse_file_num >= 30:
            high_risk_systems.append(record.app_system.name)
        system_dict[record.app_system_id] = {
            "name": record.app_system.name,
            "running_vm_num": record.running_vm_num,
            "vm_num": record.vm_num,
            "os_leak": record.os_leak_num,
            "site_leak": record.site_leak_num,
            "weak_order": record.weak_order_num,
            "horse_file": record.horse_file_num,
            "resource_use_rate": {
                "names": [],
                "cpu_use_rates": [],
                "memory_use_rates": [],
                "disk_use_rates": [],
            }
        }

    records = VirtualMachineUseRecord.objects.filter(
        time__gte=time_str_start, time__lt=time_str_end, zone=zone, app_system__isnull=False
    ).select_related("instance")
    expect_point_num = (datetime_end - datetime_start).days * 24
    init_value_list = [None] * expect_point_num
    items = ["cpu_total", "cpu_used", "memory_total", "memory_used", "disk_total", "disk_used"]
    vm_dict = {}
    for record in records:
        time_datetime = datetime.datetime.strptime(record.time, "%Y%m%d%H")
        timedelta = time_datetime - datetime_start
        index = timedelta.days * 24 + timedelta.seconds / 60 / 60
        if record.instance_id not in vm_dict:
            vm_dict[record.instance_id] = {item: copy.deepcopy(init_value_list) for item in items}
            vm_dict[record.instance_id]["vm"] = record.instance
            vm_dict[record.instance_id]["app_system_id"] = record.app_system_id
        for item in items:
            vm_dict[record.instance_id][item][index] = getattr(record, item) or None
    # print vm_dict
    for item_dict in vm_dict.values():
        app_system_id = item_dict["app_system_id"]
        if app_system_id in system_dict:
            cpu_use_rate = result_precess(list_avg(list_cal(
                item_dict.pop("cpu_used"), item_dict.pop("cpu_total"), "/", null_as="return_null"
            )) * 100)
            memory_use_rate = result_precess(list_avg(list_cal(
                item_dict.pop("memory_used"), item_dict.pop("memory_total"), "/", null_as="return_null"
            )) * 100)
            disk_use_rate = result_precess(list_avg(list_cal(
                item_dict.pop("disk_used"), item_dict.pop("disk_total"), "/", null_as="return_null"
            )) * 100)
            vm_name = item_dict["vm"].name
            system_dict[app_system_id]["resource_use_rate"]["names"].append(vm_name)
            system_dict[app_system_id]["resource_use_rate"]["cpu_use_rates"].append(cpu_use_rate)
            system_dict[app_system_id]["resource_use_rate"]["memory_use_rates"].append(memory_use_rate)
            system_dict[app_system_id]["resource_use_rate"]["disk_use_rates"].append(disk_use_rate)
    systems = list(zip(*sorted(system_dict.items()))[1]) if system_dict else []
    return [{
        "high_risk_systems": high_risk_systems,
        "systems": systems,
    }]


def download_report(request, type_, datetime_start, datetime_end, **kwargs):
    ticket_report = get_ticket_report(request, type_, datetime_start, datetime_end, **kwargs)[0]
    business_report = get_business_report(request, type_, datetime_start, datetime_end, **kwargs)[0]
    virtual_resource_report = get_virtual_resource_report(request, type_, datetime_start, datetime_end, **kwargs)[0]
    physical_resource_report = get_physical_resource_report(request, type_, datetime_start, datetime_end, **kwargs)[0]

    wb = Workbook()
    ws = wb.active

    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    # 自动换行
    wrap_text_center_align = Alignment(wrapText=True, horizontal="left", vertical="center")

    thin_style = borders.Side(border_style=borders.BORDER_THIN, color=colors.BLACK)
    medium_style = borders.Side(border_style=borders.BORDER_MEDIUM, color=colors.BLACK)
    thin_border = borders.Border(top=thin_style, bottom=thin_style, left=thin_style, right=thin_style)

    def set_cell_content(column, row, content, alignment=left_align, border=borders.DEFAULT_BORDER):
        place = "%s%s" % (column, row)
        ws[place] = content
        ws[place].font = Font(name=u"宋体", size=12)
        ws[place].alignment = alignment
        ws[place].border = border

    def set_row_content(row, contents, alignment=left_align, border=borders.DEFAULT_BORDER):
        for c_index, content in enumerate(contents, 1):
            set_cell_content(get_column_letter(c_index), row, content, alignment, border)

    ws.merge_cells("A1:D1")
    if type_ == "day":
        title = datetime_start.strftime(u"%Y-%m-%d 运维日报")
    elif type_ == "week":
        title = u"%s到%s 运维周报" % (datetime_start.strftime("%Y-%m-%d"), datetime_end.strftime("%Y-%m-%d"))
    else:
        title = datetime_start.strftime(u"%Y-%m 运维月报")
    set_cell_content("A", 1, title)
    ws["A1"].font = Font(name=u"宋体", size=16)
    ws["A1"].alignment = center_align

    ws.merge_cells("A2:D2")
    type_pretty = {
        "day": u"日",
        "week": u"周",
        "month": u"月",
    }[type_]
    summary = u"总结：\n"
    summary += u"本%s共%s个工单 ，" % (type_pretty, ticket_report["ticket_num"])
    if ticket_report["finish_rate_low_tickets"]:
        summary += u"本%s %s 完成能力弱，请注意处理积压工单。\n" % (
            type_pretty, ",".join(ticket_report["finish_rate_low_tickets"])
        )
    else:
        summary += u"无工单完成能力弱。"
    hlc_dict = physical_resource_report["high_load_cabinet"]
    if hlc_dict["cpu"] or hlc_dict["memory"] or hlc_dict["disk"]:
        summary += u"本周"
        if hlc_dict["cpu"]:
            summary += u"%s CPU使用率为高负载，" % ",".join(hlc_dict["cpu"])
        if hlc_dict["memory"]:
            summary += u"%s 内存使用率为高负载，" % ",".join(hlc_dict["memory"])
        if hlc_dict["disk"]:
            summary += u"%s 硬盘使用率为高负载，" % ",".join(hlc_dict["disk"])
        summary += u"请注意资源使用情况。\n"
    if business_report["high_risk_systems"]:
        summary += u"本%s %s 系统安全风险高，请及时防范。" % (type_, ",".join(business_report["high_risk_systems"]))
    set_cell_content("A", 2, summary, alignment=wrap_text_center_align)

    # 工单情况
    ws.merge_cells("A4:D4")
    set_cell_content("A", 4, u"工单情况", alignment=center_align)
    set_cell_content("A", 5, u"工单数量统计")
    set_row_content(6, [u"工单类型", u"工单数量", u"占比"], border=thin_border)
    current_row = 7
    tickets = ticket_report["ticket"]
    ticket_sum = sum(tickets["totals"])
    for index, ticket_name in enumerate(tickets["names"]):
        ticket_num = tickets["totals"][index]
        set_row_content(
            current_row, [ticket_name, ticket_num, result_precess(ticket_num * 100.0 / ticket_sum)], border=thin_border
        )
        current_row += 1
    current_row += 1
    set_cell_content("A", current_row, u"工单处理能力")
    current_row += 1
    set_row_content(current_row, [u"工单类型", u"完成工单数量", u"工单总量", u"完成率"], border=thin_border)
    current_row += 1
    for index, ticket_name in enumerate(tickets["names"]):
        ticket_num = tickets["totals"][index]
        finish_num = tickets["finish_nums"][index]
        finish_rate = tickets["finish_rates"][index]
        set_row_content(current_row, [ticket_name, finish_num, ticket_num, finish_rate], border=thin_border)
        current_row += 1

    # 资源情况
    current_row += 1
    ws.merge_cells("A%s:D%s" % (current_row, current_row))
    set_cell_content("A", current_row, u"资源情况", alignment=center_align)
    current_row += 1
    set_cell_content("A", current_row, u'机柜资源使用率')
    current_row += 1
    set_row_content(current_row, [u'机柜名称', u'CPU使用率', u'内存使用率', u'硬盘使用率'], border=thin_border)
    current_row += 1
    cur_dict = physical_resource_report["cabinet_use_rate"]
    for index, cabinet_name in enumerate(cur_dict["names"]):
        cpu_use_rate = cur_dict["cpu_use_rates"][index]
        memory_use_rate = cur_dict["memory_use_rates"][index]
        disk_use_rate = cur_dict["disk_use_rates"][index]
        set_row_content(current_row, [cabinet_name, cpu_use_rate, memory_use_rate, disk_use_rate], border=thin_border)
        current_row += 1

    current_row += 1
    set_cell_content("A", current_row, u'虚拟资源变化量')
    current_row += 1
    set_row_content(current_row, [u'虚拟资源类型', u'原总量', u'变化量', u'变化率'], border=thin_border)
    current_row += 1
    change_dict = virtual_resource_report["change"]
    for index, name in enumerate(change_dict["names"]):
        raw_num = change_dict["raw_nums"][index]
        num_change = change_dict["num_changes"][index]
        rate_change = change_dict["rate_changes"][index]
        set_row_content(current_row, [name, raw_num, num_change, rate_change], border=thin_border)
        current_row += 1

    current_row += 1
    set_cell_content("A", current_row, u'物理机负载趋势')
    current_row += 1
    set_row_content(current_row, [u'时间', u'CPU使用率', u'硬盘使用率', u'内存使用率'], border=thin_border)
    current_row += 1
    pm_load_trend = physical_resource_report["load_trend"]
    for index, time_ in enumerate(pm_load_trend["times"]):
        if type_ == "day":
            time_str = datetime.datetime.fromtimestamp(time_).strftime("%Y-%m-%d:%H")
        else:
            time_str = datetime.datetime.fromtimestamp(time_).strftime("%Y-%m-%d")
        cpu_use_rate = pm_load_trend["cpu_use_rates"][index]
        memory_use_rate = pm_load_trend["memory_use_rates"][index]
        disk_use_rate = pm_load_trend["disk_use_rates"][index]
        set_row_content(current_row, [time_str, cpu_use_rate, memory_use_rate, disk_use_rate], border=thin_border)
        current_row += 1

    current_row += 1
    set_cell_content("A", current_row, u'虚拟机负载趋势')
    current_row += 1
    set_row_content(current_row, [u'时间', u'CPU使用率', u'硬盘使用率', u'内存使用率'], border=thin_border)
    current_row += 1
    vm_load_trend = virtual_resource_report["load_trend"]
    for index, time_ in enumerate(vm_load_trend["times"]):
        if type_ == "day":
            time_str = datetime.datetime.fromtimestamp(time_).strftime("%Y-%m-%d:%H")
        else:
            time_str = datetime.datetime.fromtimestamp(time_).strftime("%Y-%m-%d")
        cpu_use_rate = vm_load_trend["cpu_use_rates"][index]
        memory_use_rate = vm_load_trend["memory_use_rates"][index]
        disk_use_rate = vm_load_trend["disk_use_rates"][index]
        set_row_content(current_row, [time_str, cpu_use_rate, memory_use_rate, disk_use_rate], border=thin_border)
        current_row += 1

    # 业务监控情况
    current_row += 1
    ws.merge_cells("A%s:D%s" % (current_row, current_row))
    set_cell_content("A", current_row, u"业务监控情况", alignment=center_align)
    current_row += 1
    systems = business_report["systems"]
    for system in systems:
        ws.merge_cells("A%s:D%s" % (current_row, current_row))
        set_cell_content("A", current_row, system["name"], alignment=center_align)
        current_row += 1
        set_cell_content('A', current_row, u'虚拟机使用')
        current_row += 1
        set_row_content(current_row, [u'运行中虚拟机数量', u'虚拟机总量', u'虚拟机使用率'], border=thin_border)
        current_row += 1
        running_vm_num = system['running_vm_num']
        vm_num = system['vm_num']
        set_row_content(
            current_row,
            [running_vm_num, vm_num, result_precess(running_vm_num * 100.0 / vm_num) if vm_num else "-"],
            border=thin_border
        )
        current_row += 2
        set_cell_content('A', current_row, u'安全状况')
        current_row += 1
        set_row_content(current_row, [u'安全风险类型', u'风险数量'], border=thin_border)
        current_row += 1
        risks = [
            ('os_leak', u'系统漏洞'),
            ('site_leak', u'病毒检测'),
            ('weak_order', u'弱口令'),
            ('horse_file', u'木马文件'),
        ]
        for risk, risk_name in risks:
            set_row_content(current_row, [risk_name, system[risk]], border=thin_border)
            current_row += 1
        current_row += 1
        set_cell_content('A', current_row, u'资源使用')
        current_row += 1
        set_row_content(current_row, [u'虚拟机名称', u'cpu使用率', u'内存使用率', u'硬盘使用率'], border=thin_border)
        current_row += 1
        resource_use = system['resource_use_rate']
        for index, name in enumerate(resource_use["names"]):
            cpu_use_rate = resource_use["cpu_use_rates"][index]
            memory_use_rate = resource_use["memory_use_rates"][index]
            disk_use_rate = resource_use["disk_use_rates"][index]
            set_row_content(current_row, [name, cpu_use_rate, memory_use_rate, disk_use_rate], border=thin_border)
            current_row += 1

    # 调整column宽度
    ws.column_dimensions["A"].width = 19.17
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 17.50
    ws.column_dimensions["D"].width = 21.50

    # 调整row高度
    ws.row_dimensions[1].height = 19
    ws.row_dimensions[2].height = 98
    for index in range(3, current_row):
        ws.row_dimensions[index].height = 15

    # 大边框
    def change_cell_border(cell, top=None, left=None, right=None, bottom=None):
        border = cell.border.copy()
        if top:
            border.top = top
        if left:
            border.left = left
        if right:
            border.right = right
        if bottom:
            border.bottom = bottom
        cell.border = border

    for column in ('A', 'B', 'C', 'D'):
        change_cell_border(ws['%s%s' % (column, 1)], top=medium_style)
        change_cell_border(ws['%s%s' % (column, current_row - 1)], bottom=medium_style)
    for row in range(1, current_row):
        change_cell_border(ws['A%s' % row], left=medium_style)
        change_cell_border(ws['D%s' % row], right=medium_style)
    # return wb
    return save_virtual_workbook(wb)
